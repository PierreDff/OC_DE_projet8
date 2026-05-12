"""
Consolide les fichiers Excel Weather Underground multi-onglets
en un fichier mono-onglet exploitable par Airbyte.

Chaque onglet du fichier source est nommé au format ddmmyy (ex: '011024' = 1er oct 2024).
Cette information est injectée comme colonne 'Date' dans le fichier consolidé.

Usage :
    python scripts/consolidate_wu.py

Pré-requis :
    pip install openpyxl
"""

from datetime import datetime
from pathlib import Path
import openpyxl


# Chemins en dur — adapter selon votre arborescence
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"

FILES_TO_CONSOLIDATE = [
    ("WeatherUndergroundIchtegemBE.xlsx", "WU_Ichtegem_consolidated.xlsx"),
    ("WeatherUndergroundLaMadeleineFR.xlsx", "WU_LaMadeleine_consolidated.xlsx"),
]


def parse_sheet_date(sheet_name: str) -> datetime.date:
    """Convertit un nom d'onglet 'ddmmyy' en objet date."""
    day = int(sheet_name[0:2])
    month = int(sheet_name[2:4])
    year = 2000 + int(sheet_name[4:6])
    return datetime(year, month, day).date()


def consolidate(source_path: Path, output_path: Path) -> int:
    """Consolide tous les onglets d'un .xlsx en un seul, avec colonne Date."""
    wb_src = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = "consolidated"

    header_written = False
    total_rows = 0

    for sheet_name in wb_src.sheetnames:
        sheet_date = parse_sheet_date(sheet_name)
        ws_src = wb_src[sheet_name]

        for i, row in enumerate(ws_src.iter_rows(values_only=True), start=1):
            if i == 1:  # Ligne d'entête
                if not header_written:
                    ws_dst.append(["Date"] + list(row))
                    header_written = True
                continue
            # On saute les lignes complètement vides
            if all(cell is None for cell in row):
                continue
            ws_dst.append([sheet_date.isoformat()] + list(row))
            total_rows += 1

        print(f"   ✓ Onglet '{sheet_name}' ({sheet_date}) traité")

    wb_dst.save(output_path)
    wb_src.close()
    wb_dst.close()
    return total_rows


def main():
    print(f"Répertoire de données : {DATA_DIR}\n")

    for src_name, out_name in FILES_TO_CONSOLIDATE:
        src = DATA_DIR / src_name
        out = DATA_DIR / out_name

        if not src.exists():
            print(f"⚠️  Fichier introuvable : {src}")
            continue

        print(f"=== Consolidation de {src_name} ===")
        nb = consolidate(src, out)
        print(f"   → {nb} lignes consolidées dans {out_name}\n")

    print("✅ Terminé.")


if __name__ == "__main__":
    main()
